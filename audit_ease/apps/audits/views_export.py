"""
Audit Export API Views

Endpoints:
- GET /api/v1/audits/{id}/export/csv/ - Export audit results as CSV
- GET /api/v1/audits/{id}/export/xlsx/ - Export audit results as Excel
- GET /api/v1/audits/{id}/export/pdf/ - Export audit results as PDF
"""

import csv
import logging
import io
from datetime import datetime
from django.http import StreamingHttpResponse, HttpResponse, FileResponse
from django.shortcuts import get_object_or_404
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.audits.models import Audit, Evidence
from apps.organizations.permissions import IsSameOrganization, HasActiveSubscription

# ReportLab Imports
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
except ImportError:
    pass

# OpenPyXL Imports
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, PieChart3D, Reference


logger = logging.getLogger(__name__)


class AuditExportCSVView(APIView):
    """
    Export Audit Results as CSV (Streaming Version)
    
    Optimized for large datasets using proper CSV streaming.
    """
    permission_classes = [IsAuthenticated, IsSameOrganization]

    def get(self, request, audit_id):
        try:
            # 1. Fetch Audit & Verify Permissions
            audit = get_object_or_404(Audit, id=audit_id)
            self.check_object_permissions(request, audit)
            
            # 2. Subscription Check
            if audit.organization.subscription_status != 'active':
                return Response(
                    {"error": "Subscribe to download"},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            # 3. Fetch Evidence
            evidence_list = Evidence.objects.filter(audit=audit).select_related(
                'question'
            ).order_by('created_at').values(
                'question__key',
                'question__title',
                'question__severity',
                'status',
                'raw_data',
                'comment',
                'remediation_steps'
            )
            
            logger.info(f"Streaming CSV export for audit {audit_id}")
            
            def stream_generator():
                """Yield CSV lines as a generator"""
                import io
                
                buffer = io.StringIO()
                writer = csv.writer(buffer)
                
                # Write headers
                writer.writerow([
                    'Repository',
                    'Check Name',
                    'Status',
                    'Severity',
                    'Remediation'
                ])
                yield buffer.getvalue()
                buffer.truncate(0)
                buffer.seek(0)
                
                # Write rows
                for evidence in evidence_list:
                    # Extract Resource from raw_data
                    raw = evidence.get('raw_data', {})
                    resource = "N/A"
                    if isinstance(raw, dict):
                        resource = raw.get('repo_name') or raw.get('org_name') or raw.get('name') or "N/A"

                    # Remediation text
                    remediation = evidence.get('comment', '') or evidence.get('remediation_steps', '')

                    buffer.truncate(0)
                    buffer.seek(0)
                    writer.writerow([
                        resource,
                        evidence.get('question__title', ''),
                        evidence.get('status', ''),
                        evidence.get('question__severity', ''),
                        remediation
                    ])
                    yield buffer.getvalue()
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'Audit_Report_{audit_id}.csv'
            
            response = StreamingHttpResponse(
                stream_generator(),
                content_type='text/csv'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            return response
        
        except Exception as e:
            logger.error(f"CSV Export failed: {e}")
            return Response(
                {"error": "Failed to export audit"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class AuditExportPDFView(APIView):
    """
    Export Audit Results as PDF.
    Generates a professional, industry-standard audit report.
    """
    permission_classes = [IsAuthenticated, IsSameOrganization]

    def get(self, request, audit_id):
        try:
            audit = get_object_or_404(Audit, id=audit_id)
            self.check_object_permissions(request, audit)

            if audit.organization.subscription_status != 'active':
                return Response(
                    {"error": "Subscribe to download"},
                    status=status.HTTP_403_FORBIDDEN
                )

            # Prepare Data
            from apps.audits.services.stats_service import AuditStatsService
            stats = AuditStatsService.calculate_audit_stats(audit)
            
            # Create PDF Buffer
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=letter,
                rightMargin=40, leftMargin=40,
                topMargin=40, bottomMargin=40
            )

            # Styles
            styles = getSampleStyleSheet()
            title_style = styles['Heading1']
            title_style.alignment = TA_CENTER
            
            normal_center = ParagraphStyle('NormalCenter', parent=styles['Normal'], alignment=TA_CENTER)
            
            elements = []

            # --- HEADER ---
            elements.append(Paragraph("AuditEase Security Report", title_style))
            elements.append(Spacer(1, 10))
            elements.append(Paragraph(f"Audit ID: {audit.id}", normal_center))
            elements.append(Paragraph(f"Date: {audit.created_at.strftime('%Y-%m-%d')}", normal_center))
            elements.append(Paragraph(f"Organization: {audit.organization.name}", normal_center))
            elements.append(Spacer(1, 20))

            # --- EXECUTIVE SUMMARY ---
            elements.append(Paragraph("Executive Summary", styles['Heading2']))
            elements.append(Spacer(1, 10))

            summary_data = [
                ['Metric', 'Value'],
                ['Total Checks', str(stats['total_findings'])],
                ['Failures', str(stats['failed_count'])],
                ['Passed', str(stats['passed_count'])],
                ['Compliance Score', f"{stats['pass_rate_percentage']:.1f}%"]
            ]

            summary_table = Table(summary_data, colWidths=[200, 100])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (1, 0), colors.HexColor('#003366')),
                ('TEXTCOLOR', (0, 0), (1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 25))

            # --- DETAILED FINDINGS ---
            elements.append(Paragraph("Detailed Findings (Failures Only)", styles['Heading2']))
            elements.append(Spacer(1, 10))

            # Filter for Failures only
            evidence_failures = Evidence.objects.filter(audit=audit, status='FAIL').select_related('question')

            if not evidence_failures.exists():
                elements.append(Paragraph("Great job! No failures were detected in this audit.", styles['Normal']))
            else:
                findings_data = [['Severity', 'Control', 'Resource', 'Remediation']]
                
                for ev in evidence_failures:
                    # Severity
                    sev_text = ev.question.severity
                    
                    # Resource extraction
                    raw = ev.raw_data
                    resource = "N/A"
                    if isinstance(raw, dict):
                        resource = raw.get('repo_name') or raw.get('org_name') or "N/A"
                    
                    # Remediation (Truncated if too long for table)
                    comment = ev.comment or "No details"
                    remediation = Paragraph(comment[:400] + "..." if len(comment)>400 else comment, styles['Normal'])
                    
                    findings_data.append([sev_text, ev.question.key, resource, remediation])

                # Create Table
                # Adjust column widths to fit page
                t = Table(findings_data, colWidths=[50, 80, 120, 280])
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#CCCCCC')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ]))
                
                elements.append(t)

            # --- FOOTER ---
            elements.append(Spacer(1, 40))
            elements.append(Paragraph("Generated by AuditEase - Confidential", normal_center))

            # Build PDF
            doc.build(elements)
            
            buffer.seek(0)
            return FileResponse(
                buffer, 
                as_attachment=True, 
                filename=f'Audit_Report_{audit_id}.pdf',
                content_type='application/pdf'
            )
            
        except Audit.DoesNotExist:
             return Response({"error": "Audit not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"PDF Export Error: {e}")
            return Response({"error": "Failed to generate PDF"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ExportAuditReportView(APIView):
    """
    Generate a detailed Excel report for an audit.
    Includes Executive Summary and Detailed Findings.
    """
    permission_classes = [IsAuthenticated, IsSameOrganization]

    def get(self, request, audit_id):
        try:
            audit = get_object_or_404(Audit, id=audit_id)
            self.check_object_permissions(request, audit)
            
            if audit.organization.subscription_status != 'active':
                return Response(
                    {"error": "Subscribe to download"},
                    status=status.HTTP_403_FORBIDDEN
                )

            # Dependency Check
            from apps.audits.services.stats_service import AuditStatsService

            # --- Data Aggregation ---
            stats = AuditStatsService.calculate_audit_stats(audit)
            
            total_checks = stats['total_findings']
            passed_checks = stats['passed_count']
            failed_checks = stats['failed_count']
            compliance_score = stats['pass_rate_percentage']

            # Re-query for detailed iteration (Sheet 2)
            evidence_qs = Evidence.objects.filter(audit=audit).select_related('question')

            # --- Excel Generation ---
            wb = Workbook()
            
            # 1. Sheet 1: Executive Summary
            ws_summary = wb.active
            ws_summary.title = "Executive Summary"
            
            # Title
            ws_summary.merge_cells('A1:E1')
            title_cell = ws_summary['A1']
            title_cell.value = "Audit Compliance Report"
            title_cell.font = Font(bold=True, size=16, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid") # Blue Background
            title_cell.alignment = Alignment(horizontal='center', vertical='center')

            # Summary Table
            ws_summary['A3'] = "Metric"
            ws_summary['B3'] = "Count"
            ws_summary['A4'] = "Total Checks"
            ws_summary['B4'] = total_checks
            ws_summary['A5'] = "Passed"
            ws_summary['B5'] = passed_checks
            ws_summary['A6'] = "Failed"
            ws_summary['B6'] = failed_checks
            ws_summary['A7'] = "Compliance Score"
            ws_summary['B7'] = f"{compliance_score:.1f}%"

            # Style Table Headers
            for cell in ws_summary['3']:
                cell.font = Font(bold=True)
            
            # Chart: Compliance Status
            pie = PieChart()
            pie.title = "Compliance Status"
            data = Reference(ws_summary, min_col=2, min_row=5, max_row=6) 
            labels = Reference(ws_summary, min_col=1, min_row=5, max_row=6)
            pie.add_data(data, titles_from_data=False)
            pie.set_categories(labels)
            
            # Try 3D if imported
            try:
                from openpyxl.chart import PieChart3D
                pie_3d = PieChart3D()
                pie_3d.title = "Compliance Status"
                pie_3d.add_data(data, titles_from_data=False)
                pie_3d.set_categories(labels)
                ws_summary.add_chart(pie_3d, "D3")
            except:
                ws_summary.add_chart(pie, "D3")


            # 2. Sheet 2: Detailed Findings
            ws_details = wb.create_sheet(title="Detailed Findings")
            
            headers = ["Repository", "Rule Name", "Status", "Severity", "Compliance Tag", "Remediation"]
            ws_details.append(headers)
            
            # Styling constants
            red_font = Font(color="FF0000")
            green_font = Font(color="008000")
            header_font = Font(bold=True)
            
            # Headers Styling
            for cell in ws_details[1]:
                cell.font = header_font
            
            # Freeze Top Row
            ws_details.freeze_panes = "A2"
            
            # Data Rows
            for evidence in evidence_qs:
                repo_name = evidence.question.key 
                rule_name = evidence.question.title
                status_val = evidence.status
                severity = evidence.question.severity
                compliance_tag = "SOC2"
                remediation = evidence.comment
                
                row_data = [repo_name, rule_name, status_val, severity, compliance_tag, remediation]
                ws_details.append(row_data)
                
                # Conditional Formatting
                last_row_idx = ws_details.max_row
                row_font = None
                if status_val == 'FAIL':
                    row_font = red_font
                elif status_val == 'PASS':
                    row_font = green_font
                
                if row_font:
                    for col in range(1, 7):
                        cell = ws_details.cell(row=last_row_idx, column=col)
                        cell.font = row_font

            # --- Final Response ---
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"Audit_Report_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            wb.save(response)
            return response

        except Audit.DoesNotExist:
             return Response({"error": "Audit not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"XLSX Export Error: {e}")
            return Response({"error": "Internal Server Error"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
