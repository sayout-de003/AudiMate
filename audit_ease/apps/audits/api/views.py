
import csv
import io
import logging
from datetime import datetime
from django.http import StreamingHttpResponse, HttpResponse, FileResponse
from django.shortcuts import get_object_or_404, render
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated

from apps.audits.models import Audit, Evidence
from apps.audits.serializers import AuditSerializer
from apps.organizations.permissions import IsSameOrganization
from apps.audits.services.stats_service import AuditStatsService

# ReportLab Imports
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER
except ImportError:
    pass

# OpenPyXL Imports
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# WeasyPrint & Utils
from django.template.loader import render_to_string
from utils.scoring import calculate_audit_score
try:
    from weasyprint import HTML, CSS
except ImportError:
    pass

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import base64

logger = logging.getLogger(__name__)

class AuditViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for viewing Audits and exporting results.
    """
    serializer_class = AuditSerializer
    permission_classes = [IsAuthenticated, IsSameOrganization]

    def get_queryset(self):
        """
        Filter audits by the user's organization.
        """
        if getattr(self, 'swagger_fake_view', False):
            return Audit.objects.none()
            
        organization = self.request.user.get_organization()
        return Audit.objects.filter(organization=organization).order_by('-created_at')

    @action(detail=True, methods=['get'])
    def export_csv(self, request, pk=None):
        """
        Export audit results as CSV.
        """
        try:
            # 1. Fetch Audit (IsSameOrganization handled by get_object -> get_queryset & permission_classes)
            audit = self.get_object()
            
            # 2. Subscription Check (Optional based on existing code)
            if hasattr(audit.organization, 'subscription_status') and audit.organization.subscription_status != 'active':
                 return Response(
                    {"error": "Upgrade to Premium to export data"},
                    status=status.HTTP_403_FORBIDDEN
                )

            # 3. Fetch Evidence
            evidence_list = Evidence.objects.filter(audit=audit).select_related(
                'question'
            ).order_by('created_at').values(
                'question__key',
                'question__title',
                'question__severity',
                'status',
                'raw_data',
                'comment',
                'remediation_steps'
            )
            
            logger.info(f"Streaming CSV export for audit {audit.id}")
            
            def stream_generator():
                buffer = io.StringIO()
                writer = csv.writer(buffer)
                
                # Write headers
                writer.writerow([
                    'Repository',
                    'Check Name',
                    'Status',
                    'Severity',
                    'Remediation'
                ])
                yield buffer.getvalue()
                buffer.truncate(0)
                buffer.seek(0)
                
                # Write rows
                for evidence in evidence_list:
                    # Extract Resource from raw_data
                    raw = evidence.get('raw_data', {})
                    resource = "N/A"
                    if isinstance(raw, dict):
                        resource = raw.get('repo_name') or raw.get('org_name') or raw.get('name') or "N/A"

                    # Remediation text
                    remediation = evidence.get('comment', '') or evidence.get('remediation_steps', '')

                    buffer.truncate(0)
                    buffer.seek(0)
                    writer.writerow([
                        resource,
                        evidence.get('question__title', ''),
                        evidence.get('status', ''),
                        evidence.get('question__severity', ''),
                        remediation
                    ])
                    yield buffer.getvalue()
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'Audit_Report_{audit.id}.csv'
            
            response = StreamingHttpResponse(
                stream_generator(),
                content_type='text/csv'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except Exception as e:
            logger.error(f"CSV Export failed: {e}")
            return Response(
                {"error": "Failed to export audit"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['get'])
    def export_xlsx(self, request, pk=None):
        """
        Export audit results as Excel (XLSX).
        """
        try:
            audit = self.get_object()
            
            if hasattr(audit.organization, 'subscription_status') and audit.organization.subscription_status != 'active':
                return Response(
                    {"error": "Upgrade to Premium to export data"},
                    status=status.HTTP_403_FORBIDDEN
                )

            # --- Data Aggregation ---
            stats = AuditStatsService.calculate_audit_stats(audit)
            
            total_checks = stats.get('total_findings', 0)
            passed_checks = stats.get('passed_count', 0)
            failed_checks = stats.get('failed_count', 0)
            compliance_score = stats.get('pass_rate_percentage', 0)

            # Re-query for detailed iteration
            evidence_qs = Evidence.objects.filter(audit=audit).select_related('question')

            # --- Excel Generation ---
            wb = Workbook()
            
            # 1. Sheet 1: Executive Summary
            ws_summary = wb.active
            ws_summary.title = "Executive Summary"
            
            # Title
            ws_summary.merge_cells('A1:E1')
            title_cell = ws_summary['A1']
            title_cell.value = "Audit Compliance Report"
            title_cell.font = Font(bold=True, size=16, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
            title_cell.alignment = Alignment(horizontal='center', vertical='center')

            # Summary Table
            ws_summary['A3'] = "Metric"
            ws_summary['B3'] = "Count"
            ws_summary['A4'] = "Total Checks"
            ws_summary['B4'] = total_checks
            ws_summary['A5'] = "Passed"
            ws_summary['B5'] = passed_checks
            ws_summary['A6'] = "Failed"
            ws_summary['B6'] = failed_checks
            ws_summary['A7'] = "Compliance Score"
            ws_summary['B7'] = f"{compliance_score:.1f}%"

            for cell in ws_summary['3']:
                cell.font = Font(bold=True)

            # 2. Sheet 2: Detailed Findings
            ws_details = wb.create_sheet(title="Detailed Findings")
            
            headers = ["Repository", "Rule Name", "Status", "Severity", "Compliance Tag", "Remediation"]
            ws_details.append(headers)
            
            red_font = Font(color="FF0000")
            green_font = Font(color="008000")
            header_font = Font(bold=True)
            
            for cell in ws_details[1]:
                cell.font = header_font
            
            ws_details.freeze_panes = "A2"
            
            for evidence in evidence_qs:
                # Handle raw_data safely
                raw = evidence.raw_data if isinstance(evidence.raw_data, dict) else {}
                repo_name = raw.get('repo_name') or evidence.question.key 
                
                rule_name = evidence.question.title
                status_val = evidence.status
                severity = evidence.question.severity
                compliance_tag = "SOC2" # Placeholder or derived
                remediation = evidence.comment or ""
                
                row_data = [repo_name, rule_name, status_val, severity, compliance_tag, remediation]
                ws_details.append(row_data)
                
                last_row_idx = ws_details.max_row
                row_font = None
                if status_val == 'FAIL':
                    row_font = red_font
                elif status_val == 'PASS':
                    row_font = green_font
                
                if row_font:
                    for col in range(1, 7):
                        cell = ws_details.cell(row=last_row_idx, column=col)
                        cell.font = row_font

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"Audit_Report_{audit.id}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            wb.save(response)
            return response

        except Exception as e:
            logger.error(f"XLSX Export Error: {e}")
            return Response({"error": "Internal Server Error"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'])
    def pdf(self, request, pk=None):
        """
        Export Audit results as PDF.
        """
        try:
            audit = self.get_object()
            
            if hasattr(audit.organization, 'subscription_status') and audit.organization.subscription_status != 'active':
                return Response(
                    {"error": "Upgrade to Premium to export data"},
                    status=status.HTTP_403_FORBIDDEN
                )

            # Prepare Data
            stats = AuditStatsService.calculate_audit_stats(audit)
            
            # Create PDF Buffer
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=letter,
                rightMargin=40, leftMargin=40,
                topMargin=40, bottomMargin=40
            )

            # Styles
            styles = getSampleStyleSheet()
            title_style = styles['Heading1']
            title_style.alignment = TA_CENTER
            normal_center = ParagraphStyle('NormalCenter', parent=styles['Normal'], alignment=TA_CENTER)
            
            elements = []

            # --- HEADER ---
            elements.append(Paragraph("AuditEase Security Report", title_style))
            elements.append(Spacer(1, 10))
            elements.append(Paragraph(f"Audit ID: {audit.id}", normal_center))
            elements.append(Paragraph(f"Date: {audit.created_at.strftime('%Y-%m-%d')}", normal_center))
            elements.append(Paragraph(f"Organization: {audit.organization.name}", normal_center))
            elements.append(Spacer(1, 20))

            # --- EXECUTIVE SUMMARY ---
            elements.append(Paragraph("Executive Summary", styles['Heading2']))
            elements.append(Spacer(1, 10))

            summary_data = [
                ['Metric', 'Value'],
                ['Total Checks', str(stats.get('total_findings', 0))],
                ['Failures', str(stats.get('failed_count', 0))],
                ['Passed', str(stats.get('passed_count', 0))],
                ['Compliance Score', f"{stats.get('pass_rate_percentage', 0):.1f}%"]
            ]

            summary_table = Table(summary_data, colWidths=[200, 100])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (1, 0), colors.HexColor('#003366')),
                ('TEXTCOLOR', (0, 0), (1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 25))

            # --- DETAILED FINDINGS ---
            elements.append(Paragraph("Detailed Findings (Failures Only)", styles['Heading2']))
            elements.append(Spacer(1, 10))

            # Filter for Failures only
            evidence_failures = Evidence.objects.filter(audit=audit, status='FAIL').select_related('question')

            if not evidence_failures.exists():
                elements.append(Paragraph("Great job! No failures were detected in this audit.", styles['Normal']))
            else:
                findings_data = [['Severity', 'Control', 'Resource', 'Remediation']]
                
                for ev in evidence_failures:
                    sev_text = ev.question.severity
                    
                    # Resource extraction
                    raw = ev.raw_data
                    resource = "N/A"
                    if isinstance(raw, dict):
                        resource = raw.get('repo_name') or raw.get('org_name') or "N/A"
                    
                    comment = ev.comment or "No details"
                    remediation = Paragraph(comment[:400] + "..." if len(comment)>400 else comment, styles['Normal'])
                    
                    findings_data.append([sev_text, ev.question.key, resource, remediation])

                # Create Table
                t = Table(findings_data, colWidths=[50, 80, 120, 280])
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#CCCCCC')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ]))
                
                elements.append(t)

            # --- FOOTER ---
            elements.append(Spacer(1, 40))
            elements.append(Paragraph("Generated by AuditEase - Confidential", normal_center))

            doc.build(elements)
            
            buffer.seek(0)
            return FileResponse(
                buffer, 
                as_attachment=True, 
                filename=f'Audit_Report_{audit.id}.pdf',
                content_type='application/pdf'
            )
            
        except Exception as e:
            logger.error(f"PDF Export Error: {e}")
            return Response({"error": "Failed to generate PDF"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    @action(detail=True, methods=['get'])
    def report(self, request, pk=None):
        """
        Generate a professional high-fidelity PDF report using WeasyPrint.
        """
        try:
            audit = self.get_object()
            
            # 1. Scoring
            score = calculate_audit_score(audit)
            audit.score_value = score
            audit.save(update_fields=['score_value'])

            # 2. Stats Calculation
            # We need specific counts for the template: passing, critical, high
            evidence_qs = Evidence.objects.filter(audit=audit).select_related('question')
            
            # DEBUG
            print(f"DEBUG: Audit ID: {audit.id}")
            print(f"DEBUG: Evidence Count: {evidence_qs.count()}")
            
            total_checks = evidence_qs.count()
            passed_checks = evidence_qs.filter(status='PASS').count()
            failed_checks = evidence_qs.filter(status='FAIL').count()
            
            critical_fails = evidence_qs.filter(status='FAIL', question__severity='CRITICAL').count()
            high_fails = evidence_qs.filter(status='FAIL', question__severity='HIGH').count()
            
            stats = {
                'passing': passed_checks,
                'critical': critical_fails,
                'high': high_fails,
                'failed': failed_checks,
                'total': total_checks
            }

            # 3. Construct "Checks" List & Sorting
            # Severity Map for sorting
            severity_map = {'CRITICAL': 0, 'HIGH': 1, 'MEDIUM': 2, 'LOW': 3}
            
            # Build list of wrapper objects/dicts to match template expectations
            # Template expects: check.rule_id, check.title, check.status, check.severity, check.description, check.evidence
            # Sorting: Critical -> High -> Med -> Low -> Pass
            # Interpretation: Failed items sorted by severity, then Passed items.
            
            checks = []
            for ev in evidence_qs:
                check_data = {
                    'rule_id': ev.question.key,
                    'title': ev.question.title,
                    'description': ev.question.description,
                    'severity': ev.question.severity,
                    'status': ev.status,
                    'evidence': ev, # Direct model access for screenshot, status_state
                }
                checks.append(check_data)
                
            def sort_key(c):
                # Primary sort: Status (Fail=0, Pass=1)
                status_order = 1 if c['status'] == 'PASS' else 0
                # Secondary sort: Severity index
                sev_index = severity_map.get(c['severity'], 4)
                return (status_order, sev_index)

            checks_sorted = sorted(checks, key=sort_key)

            # 4. Pie Chart (Matplotlib)
            pie_chart_base64 = None
            if passed_checks + failed_checks > 0:
                plt.figure(figsize=(6, 6))
                labels = ['Pass', 'Fail']
                sizes = [passed_checks, failed_checks]
                colors = ['#2e7d32', '#c62828'] # Green, Red
                
                plt.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%', startangle=90)
                plt.axis('equal')
                
                buffer = io.BytesIO()
                plt.savefig(buffer, format='png', bbox_inches='tight')
                buffer.seek(0)
                pie_chart_base64 = base64.b64encode(buffer.read()).decode('utf-8')
                plt.close()

            # 5. Render HTML
            context = {
                'audit': audit,
                'score': score,
                'stats': stats,
                'checks': checks_sorted, # Renamed from evidence_list to checks
                'pie_chart': pie_chart_base64,
            }
            
            html_string = render_to_string('reports/audit_report.html', context)
            
            # 6. Generate PDF
            pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()
            
            response = HttpResponse(pdf_file, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="Audit_Report_{audit.id}.pdf"'
            return response

        except Exception as e:
            logger.error(f"WeasyPrint PDF Generation Failed: {e}", exc_info=True)
            return Response(
                {"error": "Failed to generate professional report"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
